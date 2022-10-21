import os
from time import sleep

from dotenv import load_dotenv, find_dotenv
import datetime

import undetected_chromedriver as uc
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side


def login_gmail(driver, login_page, account, wait):
    search_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-social-btn_g')))
    search_btn.click()
    google_auth = driver.window_handles[2]
    driver.switch_to.window(google_auth)
    input_email = wait.until(EC.presence_of_element_located(
        (By.NAME, 'identifier')))
    input_email.send_keys(account['email'], Keys.ENTER)
    sleep(1)
    input_pwd = wait.until(EC.presence_of_element_located(
        (By.NAME, 'password')))
    input_pwd.send_keys(account['pwd'], Keys.ENTER)
    driver.switch_to.window(login_page)
    resume_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-btn_main')))
    resume_btn.click()


def login_mail(driver, login_page, account, wait):
    search_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-social-btn_mailru')))
    search_btn.click()
    mail_auth = driver.window_handles[2]
    driver.switch_to.window(mail_auth)
    input_email = wait.until(EC.presence_of_element_located(
        (By.NAME, 'username')))
    input_email.send_keys(account['email'], Keys.ENTER)
    driver.implicitly_wait(10)
    input_pwd = wait.until(EC.presence_of_element_located(
        (By.NAME, 'password')))
    input_pwd.send_keys(account['pwd'], Keys.ENTER)
    driver.switch_to.window(login_page)
    resume_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-btn_main')))
    resume_btn.click()


def login_straight(account, wait):
    input_email = wait.until(EC.presence_of_element_located(
        (By.NAME, 'login')))
    input_pwd = wait.until(EC.presence_of_element_located(
        (By.NAME, 'password')))
    input_email.send_keys(account['email'])
    input_pwd.send_keys(account['pwd'])
    login_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'gtm_login_btn')))
    login_btn.click()
    driver.implicitly_wait(10)
    resume_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-btn_main')))
    resume_btn.click()


def login_vk(driver, login_page, wait):
    search_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-social-btn_vkc')))
    search_btn.click()
    vk_auth = driver.window_handles[2]
    driver.switch_to.window(vk_auth)
    auth_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'vkuiButton')))
    auth_btn.click()
    driver.switch_to.window(login_page)
    resume_btn = wait.until(EC.presence_of_element_located(
        (By.CLASS_NAME, 'ph-btn_main')))
    resume_btn.click()


def create_new_account(account):
    wb = load_workbook('marathon.xlsx')
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == account['name']:
            account_row = row
            break
    else:
        account_cell = sheet.cell(row=sheet.max_row + 1, column=1)
        sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row + 3, end_column=1)
        account_cell.value = account['name']
        account_row = account_cell.row
        sheet.cell(row=account_row, column=2).value = 'ЛА/ДР'
        sheet.cell(row=account_row + 1, column=2).value = 'МИ'
        sheet.cell(row=account_row + 2, column=2).value = 'Морай'
        sheet.cell(row=account_row + 3, column=2).value = 'Призывной'

    wb.save('marathon.xlsx')
    wb.close()
    return account_row


def column_date():
    wb = load_workbook('marathon.xlsx')
    sheet = wb.active
    today = datetime.date.today()
    max_cell = sheet.cell(row=1, column=sheet.max_column).value
    print(max_cell)
    if max_cell is None:
        date_create(sheet, today)
    else:
        print(type(max_cell))
        if isinstance(max_cell, datetime.datetime):
            max_cell = max_cell.date()
        print(max_cell == today)
        if max_cell != today:
            date_create(sheet, today)
    wb.save('marathon.xlsx')
    wb.close()
    return sheet.max_column


def date_create(sheet, day):
    if sheet.max_column == 1:
        date_cell = sheet.cell(row=1, column=3)
    else:
        date_cell = sheet.cell(row=1, column=sheet.max_column + 1)
    date_cell.value = day
    sheet.column_dimensions[date_cell.coordinate[:1]].width = 13


def info_write(row, column, elements):
    wb = load_workbook('marathon.xlsx')
    sheet = wb.active
    for count in range(4):
        sheet.cell(row=row + count, column=column).value = elements[5 + count].text
    wb.save('marathon.xlsx')
    wb.close()


def progress(column):
    if column != 3:
        wb = load_workbook('marathon.xlsx')
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            previous = sheet.cell(row=row, column=column - 1)
            current = sheet.cell(row=row, column=column)
            if previous.value == current.value:
                current.fill = PatternFill('solid', fgColor="ff0000")
            else:
                current.fill = PatternFill('solid', fgColor="008000")
        wb.save('marathon.xlsx')
        wb.close()



def set_border():
    wb = load_workbook('marathon.xlsx')
    sheet = wb.active
    thin = Side(border_style="thin", color="000000")
    for row in sheet:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wb.save('marathon.xlsx')
    wb.close()


if not find_dotenv():
    exit('Переменные окружения не загружены т.к отсутствует файл .env')
else:
    load_dotenv()

if __name__ == '__main__':
    print(f'Time start: {datetime.datetime.now()}')
    try:
        wb = load_workbook('marathon.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Marathon'
        sheet.column_dimensions['B'].width = 15
    wb.save('marathon.xlsx')
    wb.close()
    date_column = column_date()
    count_account = int(os.getenv('TOTAL_ACCOUNTS'))
    for count in range(1, count_account + 1):
        account = {
            'name': os.getenv(f'NAME_{count}'),
            'pwd': os.getenv(f'PWD_{count}'),
            'email': os.getenv(f'EMAIL_{count}'),
            'type_auth': os.getenv(f'TYPE_{count}')
        }
        print(f'start work account {account["name"]}')
        account_row = create_new_account(account)
        options = uc.ChromeOptions()
        prefs = {"credentials_enable_service": False,
                 "profile.password_manager_enabled": False}
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-infobars')
        options.add_argument('--no-sandbox')
        options.add_argument('--start-maximized')
        options.add_experimental_option("prefs", prefs)
        if account['type_auth'] != 'gmail' and account['type_auth'] != 'vk':
            options.headless = True
        driver = uc.Chrome(use_subprocess=True, options=options)
        wait = WebDriverWait(driver, 10)
        if account['type_auth'] == 'vk':
            driver.get('https://vk.com/')
            driver.implicitly_wait(10)
            input_phone = wait.until(EC.presence_of_element_located(
                (By.CLASS_NAME, 'VkIdForm__input')))
            input_phone.send_keys(account['email'], Keys.ENTER)
            sleep(1)
            input_pwd = wait.until(EC.presence_of_element_located(
                (By.NAME, 'password')))
            input_pwd.send_keys(account['pwd'], Keys.ENTER)
            sleep(1)
            driver.get('https://pw.mail.ru/supermarathon.php')
        else:
            driver.get('https://pw.mail.ru/supermarathon.php')

        start_page = driver.current_window_handle
        login_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
            (By.CLASS_NAME, "login-button")))
        login_btn.click()
        login_page = driver.window_handles[1]
        driver.switch_to.window(login_page)
        if account['type_auth'] == 'mail':
            login_mail(driver, login_page, account, wait)
        elif account['type_auth'] == 'straight':
            login_straight(account, wait)
        elif account['type_auth'] == 'gmail':
            login_gmail(driver, login_page, account, wait)
        elif account['type_auth'] == 'vk':
            login_vk(driver, login_page, wait)

        driver.switch_to.window(start_page)
        elements = driver.find_elements(by=By.TAG_NAME, value='span')
        info_write(account_row, date_column, elements)
        driver.quit()
    progress(date_column)
    set_border()
    print('Complete!')
    print(f'Time end: {datetime.datetime.now()}')

